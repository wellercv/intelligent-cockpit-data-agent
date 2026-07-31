from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from data_insight.config import Settings
from data_insight.data_contracts import ContractRegistry
from data_insight.planner import OfflinePlanner
from data_insight.providers.composite import CompositeProvider
from data_insight.providers.governance import NLUGovernanceAdapter
from data_insight.providers.nlu import NLUEvaluationProvider
from data_insight.schemas import ConversationContext, ToolCall


def _write_nlu_report(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "总览"
    summary.append(["离线NLU黑盒测试集重测报告"])
    summary.append(["模型：fixture/model.gguf"])
    summary.append(["协议：fixture/protocol.xlsx"])
    summary.append(["测试集：fixture/jsons（2个文件）"])
    summary.append(["样本总数：10"])
    summary.append(["推理：fixture"])
    summary.append([])
    summary.append(["一、整体准确率"])
    summary.append(["口径", "说明", "正确数", "准确率"])
    summary.append(["① 原始标注", "raw", 5, "50.00%"])
    summary.append(["② + 数值槽位修正", "numeric", 6, "60.00%"])
    summary.append(["③ + 语言命名修正", "corrected", 7, "70.00%"])
    summary.append(["结论"])
    summary.append([])
    summary.append(["二、分语言准确率"])
    summary.append(["语言", "样本数", "原始标注", "修正后", "提升"])
    summary.append(["Arabic", 6, "50.00%", "66.67%", "+16.67%"])
    summary.append(["English", 4, "50.00%", "75.00%", "+25.00%"])
    summary.append([])
    summary.append(["三、分域准确率"])
    summary.append(["Domain", "样本数", "原始标注", "修正后", "提升"])
    summary.append(["carControl", 7, "42.86%", "71.43%", "+28.57%"])
    summary.append(["mediaControl", 3, "66.67%", "66.67%", "+0.00%"])

    language_issues = workbook.create_sheet("标注问题-语言命名")
    language_issues.append(["语言命名问题，共 1 条受影响样本"])
    language_issues.append(["文件", "受影响样本数", "建议"])
    language_issues.append(
        [
            "am_carControl_nlu_Arabic_0_datasets.json",
            1,
            "将 language 从 Arabic 改为 Saudi_Arabic",
        ]
    )

    numeric_issues = workbook.create_sheet("标注问题-数值槽位")
    numeric_issues.append(["数值槽位问题，共 2 条"])
    numeric_issues.append(
        ["文件", "语言", "查询语句", "Domain/Intent", "原始标注(错误)", "修正后(正确)"]
    )
    numeric_issues.append(
        [
            "am_carControl_nlu_Arabic_0_datasets.json",
            "Arabic",
            "fixture query one",
            "carControl/ADJ_VOLUME",
            json.dumps({"exact_grade": "3"}),
            json.dumps({"exact_grade": 3}),
        ]
    )
    numeric_issues.append(
        [
            "am_mediaControl_nlu_English_0_datasets.json",
            "English",
            "fixture query two",
            "mediaControl/SET_VOLUME",
            json.dumps({"ratio": "0.5"}),
            json.dumps({"ratio": 0.5}),
        ]
    )

    errors = workbook.create_sheet("模型错误明细")
    errors.append(["模型错误明细，共 3 条"])
    errors.append(["文件", "语言", "查询语句", "错误类型", "期望(修正后)", "模型预测"])

    def frame(language: str, domain: str, intent: str, slots: dict) -> str:
        return json.dumps(
            {
                "language": language,
                "nlu": [{"domain": domain, "intent": intent, "slots": slots}],
            }
        )

    errors.append(
        [
            "am_carControl_nlu_Arabic_0_datasets.json",
            "Arabic",
            "fixture error one",
            "slots",
            frame("Saudi_Arabic", "carControl", "ADJ_VOLUME", {"exact_grade": 3}),
            frame("Saudi_Arabic", "carControl", "ADJ_VOLUME", {"exact_grade": 4}),
        ]
    )
    errors.append(
        [
            "am_carControl_nlu_Arabic_0_datasets.json",
            "Arabic",
            "fixture error two",
            "intent",
            frame("Saudi_Arabic", "carControl", "OPEN_WINDOW", {}),
            frame("Saudi_Arabic", "carControl", "CLOSE_WINDOW", {}),
        ]
    )
    errors.append(
        [
            "am_mediaControl_nlu_English_0_datasets.json",
            "English",
            "fixture parse failure",
            "解析失败",
            frame("English", "mediaControl", "PLAY_SOURCE", {}),
            "not-json",
        ]
    )
    workbook.save(path)


def _settings(tmp_path: Path, report_path: Path) -> Settings:
    project_root = tmp_path / "project"
    for name in ("data", "knowledge", "skills", "config/contracts", "data/governance"):
        (project_root / name).mkdir(parents=True, exist_ok=True)
    return Settings(
        project_root=project_root,
        data_root=tmp_path,
        warehouse_path=project_root / "data" / "warehouse.duckdb",
        state_db_path=project_root / "data" / "agent_state.db",
        knowledge_dir=project_root / "knowledge",
        skills_dir=project_root / "skills",
        contracts_dir=project_root / "config" / "contracts",
        governance_dir=project_root / "data" / "governance",
        languages={"Arabic": "Arabic", "English": "English"},
        case_pattern="*_{language}/*_asr.csv",
        summary_pattern="*_{language}/*_output.json",
        nlu_report_path=report_path,
    )


@pytest.fixture
def nlu_provider(tmp_path: Path) -> NLUEvaluationProvider:
    report_path = tmp_path / "nlu-report.xlsx"
    _write_nlu_report(report_path)
    return NLUEvaluationProvider(_settings(tmp_path, report_path))


def test_nlu_provider_ingests_summary_and_detail(nlu_provider: NLUEvaluationProvider):
    assert nlu_provider.health() == {
        "provider": "nlu_evaluation",
        "ready": True,
        "optional": True,
        "synthetic_demo": False,
        "source": "nlu-report.xlsx",
        "samples": 10,
        "model_errors": 3,
        "corrected_accuracy_pct": 70.0,
        "languages": ["Arabic", "English"],
        "domains": ["carControl", "mediaControl"],
        "intents": 3,
        "source_files": 2,
        "scope": "summary plus label issues and model-error details",
    }

    overview = nlu_provider.execute(ToolCall(name="nlu_report_overview"))
    assert overview.success is True
    assert overview.data["sample_count"] == 10
    assert overview.data["corrected_correct"] == 7
    assert overview.data["model_errors"] == 3
    assert overview.data["parse_failures"] == 1


def test_nlu_provider_tools_query_report_subsets(nlu_provider: NLUEvaluationProvider):
    comparison = nlu_provider.execute(
        ToolCall(
            name="nlu_compare_accuracy",
            arguments={"dimension": "language", "order": "worst"},
        )
    )
    assert [item["language"] for item in comparison.rows] == ["Arabic", "English"]

    breakdown = nlu_provider.execute(
        ToolCall(name="nlu_error_breakdown", arguments={"group_by": "error_type"})
    )
    assert {item["error_type"] for item in breakdown.rows} == {
        "intent",
        "slots",
        "parse_failure",
    }

    search = nlu_provider.execute(
        ToolCall(
            name="search_nlu_errors",
            arguments={"error_type": "parse_failure"},
        )
    )
    assert search.data["total_matches"] == 1
    detail = nlu_provider.execute(
        ToolCall(
            name="get_nlu_error_detail",
            arguments={"error_id": search.rows[0]["error_id"]},
        )
    )
    assert detail.data["predicted_parse_ok"] is False

    quality = nlu_provider.execute(ToolCall(name="nlu_label_quality"))
    assert quality.data["numeric_label_issues"] == 2
    assert quality.data["language_naming_affected"] == 1


def test_nlu_planner_routes_analysis_and_governance_tools(
    nlu_provider: NLUEvaluationProvider,
):
    planner = OfflinePlanner(nlu_provider.tool_catalog())
    overview = planner.next_step(
        "What is the NLU corrected accuracy?",
        ConversationContext(),
        [],
        0,
    )
    assert overview.plan.calls[0].name == "nlu_report_overview"

    quality = planner.next_step(
        "Show NLU label quality issues",
        ConversationContext(),
        [],
        0,
    )
    assert quality.plan.calls[0].name == "nlu_label_quality"
    assert quality.understanding.intent == "data_governance"


def test_nlu_governance_adapter_is_read_only(
    nlu_provider: NLUEvaluationProvider,
):
    contract_root = Path(__file__).parents[1] / "config" / "contracts"
    contract = ContractRegistry(contract_root).get("nlu_evaluation")
    adapter = NLUGovernanceAdapter(nlu_provider, contract)

    findings = adapter.scan()
    assert {item.rule_id for item in findings} == {
        "LANGUAGE_NAMING_MISMATCH",
        "NUMERIC_SLOT_TYPE_MISMATCH",
        "PREDICTION_JSON_PARSE_FAILURE",
    }
    with pytest.raises(PermissionError, match="immutable evaluation artifact"):
        adapter.preview_change(findings[0].entity_key, "detail", "changed")


def test_nlu_provider_degrades_when_report_is_missing(tmp_path: Path):
    missing = tmp_path / "missing.xlsx"
    provider = NLUEvaluationProvider(_settings(tmp_path, missing))

    assert provider.health()["ready"] is False
    observation = provider.execute(ToolCall(name="nlu_report_overview"))
    assert observation.success is False
    assert "does not exist" in observation.error


def test_composite_provider_rejects_duplicate_tool_names(
    nlu_provider: NLUEvaluationProvider,
):
    with pytest.raises(ValueError, match="Duplicate tool"):
        CompositeProvider([nlu_provider, nlu_provider])


@pytest.mark.requires_business_data
def test_real_service_coordinates_asr_and_nlu(service):
    answer = service.ask(
        "Compare ASR and NLU overall accuracy",
        use_investigation_memory=False,
    )
    assert answer.tools_used == ["nlu_report_overview", "dataset_overview"]
    assert answer.grounded is True
    assert "104,897" in answer.answer_markdown
    assert "92,301" in answer.answer_markdown