"""DuckDB warehouse for the offline NLU evaluation report."""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

import duckdb
from openpyxl import load_workbook

_SCHEMA_VERSION = "1"
_REQUIRED_SHEETS = {
    "summary": "总览",
    "language_issues": "标注问题-语言命名",
    "numeric_issues": "标注问题-数值槽位",
    "model_errors": "模型错误明细",
}
_ERROR_TYPES = {
    "语言": "language",
    "解析失败": "parse_failure",
    "domain": "domain",
    "intent": "intent",
    "slots": "slots",
}


@dataclass(frozen=True)
class NLUIngestReport:
    sample_count: int
    raw_correct: int
    corrected_correct: int
    model_error_count: int
    raw_accuracy_pct: float
    corrected_accuracy_pct: float
    numeric_label_issue_count: int
    language_naming_affected: int
    parse_failure_count: int
    duplicate_label_rows: int
    duplicate_error_rows: int
    languages: list[str]
    domains: list[str]
    intent_count: int
    source_files: int
    fingerprint: str


class NLUReportWarehouse:
    def __init__(self, report_path: Path, warehouse_path: Path) -> None:
        self.report_path = report_path.resolve()
        self.warehouse_path = warehouse_path.resolve()

    def connect(self, read_only: bool = False) -> duckdb.DuckDBPyConnection:
        return duckdb.connect(str(self.warehouse_path), read_only=read_only)

    def source_fingerprint(self) -> str:
        digest = hashlib.sha256()
        digest.update(_SCHEMA_VERSION.encode("ascii"))
        with self.report_path.open("rb") as source:
            for block in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(block)
        return digest.hexdigest()

    def ensure_ready(self) -> NLUIngestReport:
        fingerprint = self.source_fingerprint()
        if self.warehouse_path.exists():
            try:
                with self.connect(read_only=True) as connection:
                    row = connection.execute(
                        "SELECT value_json FROM nlu_report_meta WHERE key='fingerprint'"
                    ).fetchone()
                    if row and json.loads(row[0]) == fingerprint:
                        return self._report(connection, fingerprint)
            except duckdb.Error:
                pass
        return self.rebuild()

    def rebuild(self) -> NLUIngestReport:
        self.warehouse_path.parent.mkdir(parents=True, exist_ok=True)
        if self.warehouse_path.exists():
            self.warehouse_path.unlink()

        workbook = load_workbook(self.report_path, read_only=True, data_only=True)
        try:
            missing = [name for name in _REQUIRED_SHEETS.values() if name not in workbook.sheetnames]
            if missing:
                raise ValueError(f"NLU report is missing worksheets: {', '.join(missing)}")

            summary = self._parse_summary(workbook[_REQUIRED_SHEETS["summary"]])
            language_issues = self._parse_language_issues(
                workbook[_REQUIRED_SHEETS["language_issues"]]
            )
            numeric_issues, duplicate_label_rows = self._parse_numeric_issues(
                workbook[_REQUIRED_SHEETS["numeric_issues"]]
            )
            model_errors, duplicate_error_rows = self._parse_model_errors(
                workbook[_REQUIRED_SHEETS["model_errors"]]
            )
        finally:
            workbook.close()

        all_label_issues = [*language_issues, *numeric_issues]
        self._reconcile_metrics(summary, model_errors)
        self._validate(summary, language_issues, numeric_issues, model_errors)
        fingerprint = self.source_fingerprint()
        metadata = {
            **summary["metadata"],
            "fingerprint": fingerprint,
            "numeric_label_issue_count": len(numeric_issues),
            "language_naming_affected": sum(
                int(item["affected_count"]) for item in language_issues
            ),
            "parse_failure_count": sum(
                item["error_type"] == "parse_failure" for item in model_errors
            ),
            "duplicate_label_rows": duplicate_label_rows,
            "duplicate_error_rows": duplicate_error_rows,
            "intent_count": len(
                {item["expected_intent"] for item in model_errors if item["expected_intent"]}
            ),
            "source_files": len({item["source_file"] for item in model_errors}),
        }

        with self.connect() as connection:
            self._create_schema(connection)
            connection.executemany(
                "INSERT INTO nlu_report_meta VALUES (?, ?)",
                [
                    (key, json.dumps(value, ensure_ascii=False))
                    for key, value in metadata.items()
                ],
            )
            connection.executemany(
                "INSERT INTO nlu_dimension_metrics VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                [
                    (
                        item["dimension"],
                        item["name"],
                        item["total"],
                        item["raw_accuracy_pct"],
                        item["corrected_accuracy_pct"],
                        item["improvement_pct"],
                        item["corrected_correct"],
                        item["model_errors"],
                    )
                    for item in summary["metrics"]
                ],
            )
            if all_label_issues:
                connection.executemany(
                    """
                    INSERT INTO nlu_label_issues VALUES (
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    )
                    """,
                    [
                        (
                            item["issue_id"],
                            item["issue_kind"],
                            item["source_file"],
                            item["language"],
                            item["query_text"],
                            item["domain"],
                            item["intent"],
                            item["changed_slot"],
                            item["original_value"],
                            item["corrected_value"],
                            item["affected_count"],
                            item["recommendation"],
                            item["source_sheet"],
                            item["source_row"],
                        )
                        for item in all_label_issues
                    ],
                )
            if model_errors:
                connection.executemany(
                    """
                    INSERT INTO nlu_model_errors VALUES (
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    )
                    """,
                    [
                        (
                            item["error_id"],
                            item["source_file"],
                            item["language"],
                            item["query_text"],
                            item["error_type"],
                            item["expected_json"],
                            item["predicted_json"],
                            item["expected_domain"],
                            item["expected_intent"],
                            item["expected_slots_json"],
                            item["predicted_language"],
                            item["predicted_domain"],
                            item["predicted_intent"],
                            item["predicted_slots_json"],
                            item["predicted_parse_ok"],
                            item["source_sheet"],
                            item["source_row"],
                        )
                        for item in model_errors
                    ],
                )
            return self._report(connection, fingerprint)

    @staticmethod
    def rows_as_dicts(cursor: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
        columns = [item[0] for item in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _report(
        self,
        connection: duckdb.DuckDBPyConnection,
        fingerprint: str,
    ) -> NLUIngestReport:
        metadata = {
            row[0]: json.loads(row[1])
            for row in connection.execute(
                "SELECT key, value_json FROM nlu_report_meta"
            ).fetchall()
        }
        dimensions = self.rows_as_dicts(
            connection.execute(
                "SELECT dimension, name FROM nlu_dimension_metrics ORDER BY dimension, name"
            )
        )
        return NLUIngestReport(
            sample_count=int(metadata["sample_count"]),
            raw_correct=int(metadata["raw_correct"]),
            corrected_correct=int(metadata["corrected_correct"]),
            model_error_count=int(metadata["model_error_count"]),
            raw_accuracy_pct=float(metadata["raw_accuracy_pct"]),
            corrected_accuracy_pct=float(metadata["corrected_accuracy_pct"]),
            numeric_label_issue_count=int(metadata["numeric_label_issue_count"]),
            language_naming_affected=int(metadata["language_naming_affected"]),
            parse_failure_count=int(metadata["parse_failure_count"]),
            duplicate_label_rows=int(metadata["duplicate_label_rows"]),
            duplicate_error_rows=int(metadata["duplicate_error_rows"]),
            languages=[
                item["name"] for item in dimensions if item["dimension"] == "language"
            ],
            domains=[
                item["name"] for item in dimensions if item["dimension"] == "domain"
            ],
            intent_count=int(metadata["intent_count"]),
            source_files=int(metadata["source_files"]),
            fingerprint=fingerprint,
        )

    @staticmethod
    def _create_schema(connection: duckdb.DuckDBPyConnection) -> None:
        connection.execute(
            """
            CREATE TABLE nlu_report_meta (
                key VARCHAR PRIMARY KEY,
                value_json VARCHAR NOT NULL
            );
            CREATE TABLE nlu_dimension_metrics (
                dimension VARCHAR NOT NULL,
                name VARCHAR NOT NULL,
                total BIGINT NOT NULL,
                raw_accuracy_pct DOUBLE NOT NULL,
                corrected_accuracy_pct DOUBLE NOT NULL,
                improvement_pct DOUBLE NOT NULL,
                corrected_correct BIGINT NOT NULL,
                model_errors BIGINT NOT NULL,
                PRIMARY KEY (dimension, name)
            );
            CREATE TABLE nlu_label_issues (
                issue_id VARCHAR PRIMARY KEY,
                issue_kind VARCHAR NOT NULL,
                source_file VARCHAR NOT NULL,
                language VARCHAR,
                query_text VARCHAR,
                domain VARCHAR,
                intent VARCHAR,
                changed_slot VARCHAR,
                original_value VARCHAR,
                corrected_value VARCHAR,
                affected_count BIGINT NOT NULL,
                recommendation VARCHAR,
                source_sheet VARCHAR NOT NULL,
                source_row BIGINT NOT NULL
            );
            CREATE TABLE nlu_model_errors (
                error_id VARCHAR PRIMARY KEY,
                source_file VARCHAR NOT NULL,
                language VARCHAR NOT NULL,
                query_text VARCHAR NOT NULL,
                error_type VARCHAR NOT NULL,
                expected_json VARCHAR NOT NULL,
                predicted_json VARCHAR NOT NULL,
                expected_domain VARCHAR,
                expected_intent VARCHAR,
                expected_slots_json VARCHAR,
                predicted_language VARCHAR,
                predicted_domain VARCHAR,
                predicted_intent VARCHAR,
                predicted_slots_json VARCHAR,
                predicted_parse_ok BOOLEAN NOT NULL,
                source_sheet VARCHAR NOT NULL,
                source_row BIGINT NOT NULL
            );
            CREATE INDEX idx_nlu_errors_language ON nlu_model_errors(language);
            CREATE INDEX idx_nlu_errors_domain ON nlu_model_errors(expected_domain);
            CREATE INDEX idx_nlu_errors_type ON nlu_model_errors(error_type);
            CREATE INDEX idx_nlu_errors_intent ON nlu_model_errors(expected_intent);
            CREATE INDEX idx_nlu_labels_kind ON nlu_label_issues(issue_kind);
            """
        )

    @classmethod
    def _parse_summary(cls, sheet) -> dict[str, Any]:
        metadata: dict[str, Any] = {
            "report_title": cls._text(sheet.cell(1, 1).value),
            "model": cls._after_colon(sheet.cell(2, 1).value),
            "protocol": cls._after_colon(sheet.cell(3, 1).value),
            "test_set": cls._after_colon(sheet.cell(4, 1).value),
            "sample_count": cls._first_integer(sheet.cell(5, 1).value),
            "inference": cls._after_colon(sheet.cell(6, 1).value),
        }
        overall_rows: list[tuple[Any, ...]] = []
        for values in sheet.iter_rows(values_only=True):
            if isinstance(values[0], str) and values[0].startswith("①"):
                overall_rows.append(values)
            elif isinstance(values[0], str) and values[0].startswith("②"):
                overall_rows.append(values)
            elif isinstance(values[0], str) and values[0].startswith("③"):
                overall_rows.append(values)
        if len(overall_rows) != 3:
            raise ValueError("NLU summary must contain three accuracy scopes")
        metadata.update(
            {
                "raw_correct": int(overall_rows[0][2]),
                "raw_accuracy_pct": cls._percent(overall_rows[0][3]),
                "numeric_corrected_correct": int(overall_rows[1][2]),
                "numeric_corrected_accuracy_pct": cls._percent(overall_rows[1][3]),
                "corrected_correct": int(overall_rows[2][2]),
                "corrected_accuracy_pct": cls._percent(overall_rows[2][3]),
            }
        )
        metadata["model_error_count"] = (
            metadata["sample_count"] - metadata["corrected_correct"]
        )

        metrics: list[dict[str, Any]] = []
        for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
            if values[0] not in {"语言", "Domain"} or values[1] != "样本数":
                continue
            dimension = "language" if values[0] == "语言" else "domain"
            for data_row in sheet.iter_rows(min_row=row_number + 1, values_only=True):
                if not data_row[0] or not isinstance(data_row[1], (int, float)):
                    break
                total = int(data_row[1])
                corrected_accuracy = cls._percent(data_row[3])
                metrics.append(
                    {
                        "dimension": dimension,
                        "name": str(data_row[0]),
                        "total": total,
                        "raw_accuracy_pct": cls._percent(data_row[2]),
                        "corrected_accuracy_pct": corrected_accuracy,
                        "improvement_pct": cls._percent(data_row[4]),
                        "corrected_correct": 0,
                        "model_errors": 0,
                    }
                )
        if not metrics:
            raise ValueError("NLU summary contains no language/domain metrics")
        return {"metadata": metadata, "metrics": metrics}

    @classmethod
    def _parse_language_issues(cls, sheet) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=3, values_only=True),
            3,
        ):
            if not values[0]:
                continue
            source_file = str(values[0])
            language = cls._language_from_file(source_file)
            domain = cls._domain_from_file(source_file)
            rows.append(
                {
                    "issue_id": cls._stable_id(
                        "NLU-LABEL",
                        "language_naming",
                        source_file,
                        row_number,
                    ),
                    "issue_kind": "language_naming",
                    "source_file": source_file,
                    "language": language,
                    "query_text": None,
                    "domain": domain,
                    "intent": None,
                    "changed_slot": "language",
                    "original_value": "Arabic",
                    "corrected_value": "Saudi_Arabic",
                    "affected_count": int(values[1]),
                    "recommendation": cls._text(values[2]),
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                }
            )
        declared = cls._first_integer(sheet.cell(1, 1).value)
        actual = sum(int(item["affected_count"]) for item in rows)
        if actual != declared:
            raise ValueError(
                f"NLU language-naming sheet declares {declared} affected samples, "
                f"found {actual}"
            )
        return rows

    @classmethod
    def _parse_numeric_issues(
        cls,
        sheet,
    ) -> tuple[list[dict[str, Any]], int]:
        rows: list[dict[str, Any]] = []
        raw_keys: list[tuple[str, ...]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=3, values_only=True),
            3,
        ):
            if not values[0]:
                continue
            raw_keys.append(tuple(cls._text(item) for item in values[:6]))
            source_file = str(values[0])
            language = str(values[1])
            query_text = str(values[2])
            domain, _, intent = str(values[3]).partition("/")
            original = json.loads(str(values[4]))
            corrected = json.loads(str(values[5]))
            changed_slots = [
                key
                for key in sorted(set(original) | set(corrected))
                if original.get(key) != corrected.get(key)
                or type(original.get(key)) is not type(corrected.get(key))
            ]
            changed_slot = ",".join(changed_slots) or "unknown"
            rows.append(
                {
                    "issue_id": cls._stable_id(
                        "NLU-LABEL",
                        source_file,
                        query_text,
                        row_number,
                    ),
                    "issue_kind": "numeric_slot_type",
                    "source_file": source_file,
                    "language": language,
                    "query_text": query_text,
                    "domain": domain,
                    "intent": intent,
                    "changed_slot": changed_slot,
                    "original_value": json.dumps(original, ensure_ascii=False),
                    "corrected_value": json.dumps(corrected, ensure_ascii=False),
                    "affected_count": 1,
                    "recommendation": "Use protocol numeric types for numeric slots.",
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                }
            )
        declared = cls._first_integer(sheet.cell(1, 1).value)
        if len(rows) != declared:
            raise ValueError(
                f"NLU numeric-label sheet declares {declared} rows, found {len(rows)}"
            )
        return rows, len(raw_keys) - len(set(raw_keys))

    @classmethod
    def _parse_model_errors(
        cls,
        sheet,
    ) -> tuple[list[dict[str, Any]], int]:
        rows: list[dict[str, Any]] = []
        raw_keys: list[tuple[str, ...]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=3, values_only=True),
            3,
        ):
            if not values[0]:
                continue
            raw_keys.append(tuple(cls._text(item) for item in values[:6]))
            source_file = str(values[0])
            language = str(values[1])
            query_text = str(values[2])
            error_type = _ERROR_TYPES.get(str(values[3]), str(values[3]))
            expected_raw = str(values[4])
            predicted_raw = str(values[5])
            expected = json.loads(expected_raw)
            expected_frame = cls._first_frame(expected)
            try:
                predicted = json.loads(predicted_raw)
                predicted_parse_ok = True
                predicted_frame = cls._first_frame(predicted)
                predicted_language = cls._text(predicted.get("language")) or None
            except (json.JSONDecodeError, TypeError):
                predicted_parse_ok = False
                predicted_frame = {}
                predicted_language = None
            rows.append(
                {
                    "error_id": cls._stable_id(
                        "NLU-ERR",
                        source_file,
                        query_text,
                        expected_raw,
                        predicted_raw,
                        row_number,
                    ),
                    "source_file": source_file,
                    "language": language,
                    "query_text": query_text,
                    "error_type": error_type,
                    "expected_json": expected_raw,
                    "predicted_json": predicted_raw,
                    "expected_domain": expected_frame.get("domain"),
                    "expected_intent": expected_frame.get("intent"),
                    "expected_slots_json": json.dumps(
                        expected_frame.get("slots") or {},
                        ensure_ascii=False,
                    ),
                    "predicted_language": predicted_language,
                    "predicted_domain": predicted_frame.get("domain"),
                    "predicted_intent": predicted_frame.get("intent"),
                    "predicted_slots_json": json.dumps(
                        predicted_frame.get("slots") or {},
                        ensure_ascii=False,
                    ),
                    "predicted_parse_ok": predicted_parse_ok,
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                }
            )
        declared = cls._first_integer(sheet.cell(1, 1).value)
        if len(rows) != declared:
            raise ValueError(
                f"NLU model-error sheet declares {declared} rows, found {len(rows)}"
            )
        return rows, len(raw_keys) - len(set(raw_keys))

    @staticmethod
    def _validate(
        summary: dict[str, Any],
        language_issues: Sequence[dict[str, Any]],
        numeric_issues: Sequence[dict[str, Any]],
        model_errors: Sequence[dict[str, Any]],
    ) -> None:
        metadata = summary["metadata"]
        metrics = summary["metrics"]
        for dimension in ("language", "domain"):
            total = sum(
                item["total"] for item in metrics if item["dimension"] == dimension
            )
            if total != metadata["sample_count"]:
                raise ValueError(
                    f"NLU {dimension} totals {total} do not match sample count "
                    f"{metadata['sample_count']}"
                )
        if metadata["corrected_correct"] + len(model_errors) != metadata["sample_count"]:
            raise ValueError("Corrected NLU correct count plus error rows is inconsistent")
        if any(
            item["original_value"] == item["corrected_value"]
            for item in numeric_issues
        ):
            raise ValueError("NLU numeric issue contains an unchanged label")
        if sum(int(item["affected_count"]) for item in language_issues) < (
            metadata["corrected_correct"] - metadata["numeric_corrected_correct"]
        ):
            raise ValueError("NLU language naming impact is smaller than corrected gains")


    @staticmethod
    def _reconcile_metrics(
        summary: dict[str, Any],
        model_errors: Sequence[dict[str, Any]],
    ) -> None:
        errors_by_language = Counter(item["language"] for item in model_errors)
        errors_by_domain = Counter(item["expected_domain"] for item in model_errors)
        for item in summary["metrics"]:
            errors = (
                errors_by_language[item["name"]]
                if item["dimension"] == "language"
                else errors_by_domain[item["name"]]
            )
            item["model_errors"] = errors
            item["corrected_correct"] = item["total"] - errors
            calculated = round(100.0 * item["corrected_correct"] / item["total"], 2)
            if calculated != item["corrected_accuracy_pct"]:
                raise ValueError(
                    f"NLU {item['dimension']} `{item['name']}` corrected accuracy "
                    f"{calculated} does not match summary {item['corrected_accuracy_pct']}"
                )

    @staticmethod
    def _first_frame(payload: dict[str, Any]) -> dict[str, Any]:
        frames = payload.get("nlu") or []
        return frames[0] if frames and isinstance(frames[0], dict) else {}

    @staticmethod
    def _stable_id(prefix: str, *values: Any) -> str:
        payload = "\x1f".join(str(value) for value in values)
        digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:12]
        return f"{prefix}-{digest}"

    @staticmethod
    def _language_from_file(source_file: str) -> str | None:
        match = re.search(r"_nlu_([^_]+)_", source_file)
        return match.group(1) if match else None

    @staticmethod
    def _domain_from_file(source_file: str) -> str | None:
        match = re.search(r"am_(.+?)_nlu_", source_file)
        return match.group(1) if match else None

    @staticmethod
    def _after_colon(value: Any) -> str:
        text = NLUReportWarehouse._text(value)
        return re.split(r"[：:]", text, maxsplit=1)[-1].strip()

    @staticmethod
    def _first_integer(value: Any) -> int:
        match = re.search(r"\d+", NLUReportWarehouse._text(value))
        if not match:
            raise ValueError(f"Expected an integer in: {value}")
        return int(match.group(0))

    @staticmethod
    def _percent(value: Any) -> float:
        return float(NLUReportWarehouse._text(value).strip().rstrip("%+"))

    @staticmethod
    def _text(value: Any) -> str:
        return "" if value is None else str(value)