"""Deterministic synthetic data for evaluating a fresh public clone."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

DEMO_LANGUAGES = {
    "Arabic": "Arabic",
    "English": "English",
    "French": "French",
    "German": "German",
    "Italian": "Italian",
    "Portuguese": "Portuguese",
    "Spanish": "Spainsh",
}
DEMO_DOMAINS = (
    "carControl",
    "generalControl",
    "mediaControl",
    "naviControl",
    "phone",
    "systemControl",
)
_DEMO_ASR_ERRORS = {
    ("Arabic", "carControl"),
    ("English", "phone"),
    ("French", "generalControl"),
    ("German", "systemControl"),
    ("Italian", "naviControl"),
    ("Portuguese", "mediaControl"),
}


def ensure_demo_data(root: Path) -> dict[str, Any]:
    """Create a small, explicit synthetic ASR/NLU dataset once."""

    root.mkdir(parents=True, exist_ok=True)
    marker = root / ".demo-data-v1.json"
    nlu_report = root / "demo_nlu_report.xlsx"
    if marker.exists() and nlu_report.exists():
        return json.loads(marker.read_text(encoding="utf-8"))

    asr_cases = _write_asr_sources(root)
    _write_nlu_report(nlu_report)
    payload = {
        "synthetic_demo": True,
        "asr_cases": asr_cases,
        "nlu_samples": 14,
        "nlu_model_errors": 3,
        "nlu_report": str(nlu_report),
    }
    marker.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return payload


def _write_asr_sources(root: Path) -> int:
    cases = 0
    for language, folder in DEMO_LANGUAGES.items():
        for domain in DEMO_DOMAINS:
            source_dir = root / folder / f"{domain}_{language}"
            source_dir.mkdir(parents=True, exist_ok=True)
            is_error = (language, domain) in _DEMO_ASR_ERRORS
            result = "✗" if is_error else "✓"
            reference = f"{language} {domain} demo request"
            hypothesis = (
                f"{language} {domain} demo mismatch" if is_error else reference
            )
            csv_path = source_dir / f"demo_{language}_asr.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(
                    ["NO（序号）", "Result（结果）", "REF（参考）", "HYP（假设）"]
                )
                writer.writerow(["ASR #1/1", result, reference, hypothesis])
            summary = {
                "generated_at": "2026-01-01T00:00:00Z",
                "languages": {
                    language: {
                        "asr": {
                            "total": 1,
                            "correct": 0 if is_error else 1,
                            "accuracy_pct": 0.0 if is_error else 100.0,
                            "expected_total": 1,
                            "total_matched": True,
                        }
                    }
                },
            }
            (source_dir / f"demo_{language}_output.json").write_text(
                json.dumps(summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            cases += 1
    return cases


def _write_nlu_report(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "总览"
    summary.append(["Synthetic Demo NLU Re-evaluation Report"])
    summary.append(["模型：synthetic/demo-model"])
    summary.append(["协议：synthetic/demo-protocol"])
    summary.append(["测试集：generated demo data (7 languages x 6 domains)"])
    summary.append(["样本总数：14"])
    summary.append(["推理：deterministic synthetic fixture"])
    summary.append([])
    summary.append(["一、整体准确率"])
    summary.append(["口径", "说明", "正确数", "准确率"])
    summary.append(["① 原始标注", "synthetic raw", 9, "64.29%"])
    summary.append(["② + 数值槽位修正", "synthetic numeric", 10, "71.43%"])
    summary.append(["③ + 语言命名修正", "synthetic corrected", 11, "78.57%"])
    summary.append(["结论：仅用于演示 Provider 接口，不代表真实模型质量。"])
    summary.append([])
    summary.append(["二、分语言准确率"])
    summary.append(["语言", "样本数", "原始标注", "修正后", "提升"])
    language_errors = {"Arabic", "French", "Portuguese"}
    for language in DEMO_LANGUAGES:
        corrected = "50.00%" if language in language_errors else "100.00%"
        raw = "0.00%" if language == "Arabic" else corrected
        improvement = "+50.00%" if language == "Arabic" else "+0.00%"
        summary.append([language, 2, raw, corrected, improvement])
    summary.append([])
    summary.append(["三、分域准确率"])
    summary.append(["Domain", "样本数", "原始标注", "修正后", "提升"])
    domain_metrics = {
        "carControl": (3, "33.33%", "66.67%", "+33.34%"),
        "generalControl": (3, "66.67%", "66.67%", "+0.00%"),
        "mediaControl": (2, "50.00%", "50.00%", "+0.00%"),
        "naviControl": (2, "100.00%", "100.00%", "+0.00%"),
        "phone": (2, "100.00%", "100.00%", "+0.00%"),
        "systemControl": (2, "100.00%", "100.00%", "+0.00%"),
    }
    for domain, values in domain_metrics.items():
        summary.append([domain, *values])

    language_issues = workbook.create_sheet("标注问题-语言命名")
    language_issues.append(["Synthetic language naming issue, 共 1 条受影响样本"])
    language_issues.append(["文件", "受影响样本数", "建议"])
    language_issues.append(
        [
            "demo_carControl_nlu_Arabic.json",
            1,
            "Synthetic example: normalize Arabic to Saudi_Arabic",
        ]
    )

    numeric_issues = workbook.create_sheet("标注问题-数值槽位")
    numeric_issues.append(["Synthetic numeric slot issue, 共 1 条"])
    numeric_issues.append(
        ["文件", "语言", "查询语句", "Domain/Intent", "原始标注(错误)", "修正后(正确)"]
    )
    numeric_issues.append(
        [
            "demo_carControl_nlu_Arabic.json",
            "Arabic",
            "synthetic volume request",
            "carControl/ADJ_VOLUME",
            json.dumps({"exact_grade": "3"}),
            json.dumps({"exact_grade": 3}),
        ]
    )

    errors = workbook.create_sheet("模型错误明细")
    errors.append(["Synthetic model error details, 共 3 条"])
    errors.append(["文件", "语言", "查询语句", "错误类型", "期望(修正后)", "模型预测"])
    errors.append(
        [
            "demo_carControl_nlu_Arabic.json",
            "Arabic",
            "synthetic slot mismatch",
            "slots",
            _frame("Saudi_Arabic", "carControl", "ADJ_VOLUME", {"exact_grade": 3}),
            _frame("Saudi_Arabic", "carControl", "ADJ_VOLUME", {"exact_grade": 4}),
        ]
    )
    errors.append(
        [
            "demo_generalControl_nlu_French.json",
            "French",
            "synthetic intent mismatch",
            "intent",
            _frame("French", "generalControl", "ADJ_SIZE", {}),
            _frame("French", "generalControl", "SET_SIZE", {}),
        ]
    )
    errors.append(
        [
            "demo_mediaControl_nlu_Portuguese.json",
            "Portuguese",
            "synthetic parse failure",
            "解析失败",
            _frame("Portuguese", "mediaControl", "PLAY_SOURCE", {}),
            "not-json",
        ]
    )
    workbook.save(path)


def _frame(language: str, domain: str, intent: str, slots: dict[str, Any]) -> str:
    return json.dumps(
        {
            "language": language,
            "nlu": [{"domain": domain, "intent": intent, "slots": slots}],
        },
        ensure_ascii=False,
    )
